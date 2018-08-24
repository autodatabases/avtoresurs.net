from django.core.mail import EmailMessage
from django.db import transaction

from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from django.urls import reverse
from io import StringIO, BytesIO
from django.views.generic import TemplateView
from django.core.mail import send_mail
from django.contrib.auth.models import User
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter

from avtoresurs_new.settings import EMAIL_NOREPLY, EMAIL_BCC, EMAIL_NOREPLY_LIST
from avtoresurs_new.settings import EMAIL_TO
from postman.models import Message, STATUS_ACCEPTED
from profile.models import Profile
from shop.models.cart import Cart
from shop.models.order import Order, OrderProduct
from openpyxl import Workbook

from shop.views.cart import make_cart_storages


def as_text(value):
    if value is None:
        return ""
    return str(value)


def order_notification(cart, order, user):
    cart_storages = make_cart_storages(cart, user)

    for cart_storage, items in cart_storages.items():
        body = 'Новый заказ от {date} #{order_number}.\r\n\r\nИнформация о заказе:\r\n'.format(
            date=order.added.strftime('%d.%m.%Y %H:%M'), order_number=order.id)
        profile = Profile.objects.get(user=user)
        body += 'Заказчик: {name} \r\n'.format(name=profile.fullname or profile)
        body += 'Код заказчика: {vip_code}\r\n'.format(vip_code=profile.vip_code or 'код заказчика отсутствует')

        storage = cart_storage
        order_total = 0
        for idx, item in enumerate(items):
            product = item.item
            price = product.get_price(user, item.storage)
            qty = item.quantity
            order_total += price * qty
            body += '{idx}. {title} {brand} {sku}, {qty} шт. x {price} руб.., на общую сумму: {line_total} руб.\r\n'.format(
                idx=idx + 1,
                title=product.title(),
                brand=product.brand,
                sku=product.get_sku(),
                qty=qty,
                price=price,
                line_total=qty * price)
            op = OrderProduct(order=order, item=product, qty=qty, price=price)
            op.save()

        body += '\r\nИтого: {total} руб.'.format(total=order_total)
        body += '\r\nСклад: {storage}'.format(storage=storage.name)
        body += '\r\n\r\nEmail клиента: {email}'.format(email=user.email)
        body += '\r\n\r\nКомментарии к заказу:\r\n{comment}'.format(comment=order.comment)

        subject = 'Сформирован новый заказ от {date} № {order_num}!'.format(date=order.added.strftime("%d.%m.%Y %H:%M"),
                                                                            order_num=order.id)
        sender = User.objects.filter(username='admin').first()
        recipient = user
        status = STATUS_ACCEPTED
        message_user = Message(subject=subject, body=body, sender=sender, recipient=recipient, moderation_status=status)
        message_user.save()
        message_admin = Message(subject=subject, body=body, sender=recipient, recipient=sender,
                                moderation_status=status)
        message_admin.save()

        # XLS document
        wb = Workbook()
        ws = wb.active
        ws.title = 'Заказ'

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        data = []

        for idx, item in enumerate(cart.cartitem_set.all()):
            product = item.item
            price = product.get_price(user, item.storage)
            print(item.storage)
            print(price)
            qty = item.quantity
            coordinate = 'B' + str(idx)
            article = '%s %s' % (product.brand, product.get_sku())
            data.append([idx + 1, product.title(), article, qty, price, qty * price])

        ws.append(['#', 'Товар', 'Брэнд', 'Количество', 'Цена за единицу', 'Цена (общая)'])
        for row in data:
            ws.append(row)
        ws.append(['', '', '', '', 'Итого:', order_total])

        row_count = ws.max_row + 1
        column_count = ws.max_column + 1

        for i in range(1, row_count):
            for j in range(1, column_count):
                ws.cell(row=i, column=j).border = thin_border

        ws.column_dimensions["A"].width = 5.0
        ws.column_dimensions["B"].width = 17.0
        ws.column_dimensions["C"].width = 23.0
        ws.column_dimensions["D"].width = 13.0
        ws.column_dimensions["E"].width = 18.0
        ws.column_dimensions["F"].width = 18.0

        email = EmailMessage(
            subject,
            body,
            EMAIL_NOREPLY,
            [storage.email],
            EMAIL_BCC,
            reply_to=EMAIL_NOREPLY_LIST,
            headers={'Message-ID': 'foo'},
        )

        output = BytesIO()
        wb.save(output)

        email.attach('order.xlsx', output.getvalue(), 'application/excel')
        email.send()

    return True


class CheckoutView(TemplateView):
    model = Cart
    template_name = "shop/checkout_view.html"

    def get_object(self, *args, **kwargs):
        if self.request.method == 'get':
            cart_id = self.request.session.get("cart_id", None)
        else:
            cart_id = self.request.session.post("cart_id", None)
        if not cart_id:
            return redirect("cart")
        cart = Cart.objects.get(id=cart_id)

        return cart

    def get_context_data(self, *args, **kwargs):
        context = super(CheckoutView, self).get_context_data(**kwargs)
        user_checkout = None
        if self.request.user.is_authenticated():
            cart = self.get_object()
            cart_storages = make_cart_storages(cart, self.request.user)
            context["cart"] = self.get_object()
            context["cart_storages"] = cart_storages
        return context

    @transaction.atomic
    def post(self, request):
        cart = self.get_object()
        self.request.session['cart_id'] = None
        comment = request.POST.get('comment', None)
        order = Order(user=cart.user)
        order.comment = comment
        order.order_total = cart.subtotal
        order.save()
        order_notification(cart=cart, order=order, user=self.request.user)


        return HttpResponseRedirect('/checkout/success/')


class CheckoutSuccessView(TemplateView):
    template_name = "shop/checkout_success_view.html"
